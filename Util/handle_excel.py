import openpyxl
import sys
import os
base_path = os.getcwd()
sys.path.append(base_path)

class HandExcel:
    def __init__(self,filename):
        self.file_name = filename
    def load_excel(self):
        """
        加载excel
        """
        open_excel = openpyxl.load_workbook(base_path+"/Cases/"+self.file_name)
        return open_excel

    def get_sheet_data(self,index=None):
        """
        加载所有sheet
        """
        sheet_name = self.load_excel().sheetnames
        #print(sheet_name)
        if index == None:
            index = 0
        data = self.load_excel()[sheet_name[index]]
        return data

    def get_cell_value(self,row,clos):
        """
        获取某一单元格内容
        """
        data = self.get_sheet_data().cell(row=row,column=clos).value
        return data

    def get_rows(self):
        """
        获取行数
        """
        row = self.get_sheet_data().max_row
        return row
    
    def get_rows_value(self,row):
        """
        获取某一行的内容
        """
        row_list = []
        for i in self.get_sheet_data()[row]:
            row_list.append(i.value)
        return row_list

# excel_data = HandExcel(filename)


if __name__ == "__main__":
    handle = HandExcel("test.xlsx")
    print(handle.load_excel())
    print(handle.get_cell_value(2,5))
    print(handle.get_rows_value(2))
    print(handle.get_sheet_data())